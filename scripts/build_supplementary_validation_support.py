#!/usr/bin/env python3
from __future__ import annotations

import csv
import math
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from scipy import stats


ROOT = Path(__file__).resolve().parents[1]
VALIDATION_DIR = ROOT / "results" / "supplementary_validation"
SCORE_DIR = ROOT / "results" / "scores"
SC_DIR = ROOT / "results" / "singlecell"
FIG_DIR = ROOT / "results" / "figures"
PLOT_DIR = ROOT / "plots" / "publication"
SUP_DIR = ROOT / "supplementary_tables"


def read_genes(path: Path) -> set[str]:
    return {line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()}


def bh_adjust(pvalues: list[float]) -> list[float]:
    indexed = sorted((p, i) for i, p in enumerate(pvalues))
    out = [math.nan] * len(pvalues)
    prev = 1.0
    m = len(indexed)
    for rank_from_end, (p, i) in enumerate(reversed(indexed), start=1):
        rank = m - rank_from_end + 1
        adj = min(prev, p * m / rank)
        out[i] = min(adj, 1.0)
        prev = adj
    return out


def write_tsv(path: Path, rows: list[dict]) -> None:
    if not rows:
        raise ValueError(f"No rows to write for {path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()), delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)


def build_gene_set_overlap() -> pd.DataFrame:
    fixed = read_genes(SCORE_DIR / "strict_nonresponse_module_genes.txt")
    rows = []
    for holdout in ["GSE12251", "GSE16879", "GSE23597"]:
        lodo = read_genes(SCORE_DIR / "lodo_modules" / f"lodo_excluding_{holdout}_genes.txt")
        overlap = sorted(fixed & lodo)
        union = fixed | lodo
        rows.append(
            {
                "held_out_dataset": holdout,
                "fixed_consensus_n": len(fixed),
                "lodo_gene_set_n": len(lodo),
                "overlap_n": len(overlap),
                "fixed_genes_present_in_lodo_pct": len(overlap) / len(fixed) if fixed else math.nan,
                "lodo_genes_present_in_fixed_pct": len(overlap) / len(lodo) if lodo else math.nan,
                "jaccard_index": len(overlap) / len(union) if union else math.nan,
                "overlap_genes": "|".join(overlap),
            }
        )
    df = pd.DataFrame(rows)
    df.to_csv(VALIDATION_DIR / "gene_set_overlap.tsv", sep="\t", index=False)
    return df


def global_kruskal(df: pd.DataFrame, group_col: str, value_col: str, reference: str, subset: str) -> dict:
    groups = []
    for _, sub in df.groupby(group_col):
        vals = sub[value_col].dropna().to_numpy(dtype=float)
        if len(vals) > 0:
            groups.append(vals)
    stat, pvalue = stats.kruskal(*groups)
    return {
        "reference": reference,
        "subset": subset,
        "test": "Kruskal-Wallis",
        "grouping": group_col,
        "n_groups": len(groups),
        "statistic": stat,
        "pvalue": pvalue,
    }


def one_vs_rest_tests(df: pd.DataFrame, group_col: str, value_col: str, reference: str, subset: str) -> list[dict]:
    rows = []
    for group, sub in df.groupby(group_col):
        group_vals = sub[value_col].dropna().to_numpy(dtype=float)
        rest_vals = df.loc[df[group_col] != group, value_col].dropna().to_numpy(dtype=float)
        if len(group_vals) == 0 or len(rest_vals) == 0:
            continue
        stat, pvalue = stats.mannwhitneyu(group_vals, rest_vals, alternative="two-sided")
        group_mean = float(np.mean(group_vals))
        rest_mean = float(np.mean(rest_vals))
        rows.append(
            {
                "reference": reference,
                "subset": subset,
                "test": "Mann-Whitney U one-vs-rest",
                "grouping": group_col,
                "group": group,
                "comparator": "all other groups",
                "n_group": len(group_vals),
                "n_comparator": len(rest_vals),
                "group_mean": group_mean,
                "comparator_mean": rest_mean,
                "direction": "higher" if group_mean > rest_mean else "lower",
                "statistic": stat,
                "pvalue": pvalue,
            }
        )
    return rows


def build_singlecell_tests() -> tuple[pd.DataFrame, pd.DataFrame]:
    epithelial = pd.read_csv(SC_DIR / "GSE116222_epithelial_state_assignments.tsv", sep="\t")
    epithelial = epithelial[epithelial["assigned_state"] != "Unassigned"].copy()
    global_rows = [
        global_kruskal(epithelial, "assigned_state", "module_score", "GSE116222", "epithelial_states")
    ]
    pairwise_rows = one_vs_rest_tests(
        epithelial,
        "assigned_state",
        "module_score",
        "GSE116222",
        "epithelial_states",
    )

    immune = pd.read_csv(SC_DIR / "GSE125527_rectal_immune_module_scores.tsv", sep="\t")
    for disease, sub in immune.groupby("disease_assignment"):
        global_rows.append(
            global_kruskal(sub, "celltype", "module_score", "GSE125527", f"rectum_{disease}")
        )
        pairwise_rows.extend(
            one_vs_rest_tests(sub, "celltype", "module_score", "GSE125527", f"rectum_{disease}")
        )

    adjusted = bh_adjust([float(row["pvalue"]) for row in pairwise_rows])
    for row, p_adj in zip(pairwise_rows, adjusted):
        row["p_adj_bh"] = p_adj

    global_df = pd.DataFrame(global_rows)
    pairwise_df = pd.DataFrame(pairwise_rows)
    global_df.to_csv(VALIDATION_DIR / "singlecell_global_tests.tsv", sep="\t", index=False)
    pairwise_df.to_csv(VALIDATION_DIR / "singlecell_pairwise_tests.tsv", sep="\t", index=False)
    return global_df, pairwise_df


def gene_zscore(matrix: pd.DataFrame) -> pd.DataFrame:
    values = matrix.set_index("gene").astype(float)
    centered = values.sub(values.mean(axis=1), axis=0)
    sd = values.std(axis=1).replace(0, 1)
    return centered.div(sd, axis=0)


def select_top_variable(matrix: pd.DataFrame, n: int = 30) -> pd.DataFrame:
    values = matrix.set_index("gene").astype(float)
    spread = values.max(axis=1) - values.min(axis=1)
    keep = spread.sort_values(ascending=False).head(min(n, len(spread))).index
    return matrix[matrix["gene"].isin(keep)].copy()


def build_heatmap_figure(epi_means: pd.DataFrame, immune_means: pd.DataFrame) -> None:
    PLOT_DIR.mkdir(parents=True, exist_ok=True)
    epi_top = select_top_variable(epi_means, 30)
    immune_top = select_top_variable(immune_means, 30)
    epi_z = gene_zscore(epi_top)
    immune_z = gene_zscore(immune_top)

    fig, axes = plt.subplots(1, 2, figsize=(10.5, 8.0), constrained_layout=True)
    for ax, zmat, title in [
        (axes[0], epi_z, "A. Epithelial reference (GSE116222)"),
        (axes[1], immune_z, "B. Rectal immune reference (GSE125527)"),
    ]:
        im = ax.imshow(zmat.to_numpy(), aspect="auto", cmap="RdBu_r", vmin=-2.0, vmax=2.0)
        ax.set_title(title, fontsize=10, fontweight="bold")
        ax.set_xticks(np.arange(zmat.shape[1]))
        ax.set_xticklabels(zmat.columns, rotation=45, ha="right", fontsize=7)
        ax.set_yticks(np.arange(zmat.shape[0]))
        ax.set_yticklabels(zmat.index, fontsize=6)
        ax.tick_params(length=0)
        ax.set_xlabel("Cell population", fontsize=8)
        ax.set_ylabel("Signature genes", fontsize=8)

    cbar = fig.colorbar(im, ax=axes, shrink=0.65, pad=0.02)
    cbar.set_label("Row z-score of mean log-normalized expression", fontsize=8)
    for label in cbar.ax.get_yticklabels():
        label.set_fontsize(7)

    fig.suptitle("Supplementary Figure S3. Expression pattern of representative consensus-signature genes", fontsize=11, fontweight="bold")
    fig.savefig(PLOT_DIR / "FigureS3_singlecell_signature_heatmaps.png", dpi=600, bbox_inches="tight")
    fig.savefig(PLOT_DIR / "FigureS3_singlecell_signature_heatmaps.pdf", bbox_inches="tight")
    plt.close(fig)


def build_integrated_figure4_anchor() -> pd.DataFrame:
    epi = pd.read_csv(FIG_DIR / "fig4_cellstate_localization.tsv", sep="\t")
    epi.insert(0, "reference_type", "epithelial")
    epi.insert(1, "dataset_id", "GSE116222")
    epi.insert(2, "disease_assignment", "NA")

    immune = pd.read_csv(FIG_DIR / "fig4b_immune_localization.tsv", sep="\t")
    immune.insert(0, "reference_type", "immune_rectum")

    cols = [
        "reference_type",
        "dataset_id",
        "disease_assignment",
        "cell_state",
        "enrichment_score",
        "median_module_score",
        "n_cells",
        "present_module_genes",
        "supporting_genes",
        "interpretation",
    ]
    integrated = pd.concat([epi[cols], immune[cols]], ignore_index=True)
    integrated.to_csv(FIG_DIR / "fig4_localization_integrated.tsv", sep="\t", index=False)
    return integrated


def write_workbook(
    overlap: pd.DataFrame,
    global_tests: pd.DataFrame,
    pairwise_tests: pd.DataFrame,
    epi_means: pd.DataFrame,
    immune_means: pd.DataFrame,
) -> Path:
    SUP_DIR.mkdir(parents=True, exist_ok=True)
    path = SUP_DIR / "TableS7_gene_set_overlap_and_singlecell_tests.xlsx"
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    sheets = {
        "GeneSetOverlap": overlap,
        "SingleCellGlobalTests": global_tests,
        "SingleCellPairwiseTests": pairwise_tests,
        "EpithelialGeneMeans": epi_means,
        "ImmuneGeneMeans": immune_means,
    }
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for name, df in sheets.items():
        ws = wb.create_sheet(name)
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))
        for cell in ws[1]:
            cell.font = Font(name="Arial", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for col in ws.columns:
            width = min(max(len(str(cell.value)) if cell.value is not None else 0 for cell in col) + 2, 45)
            ws.column_dimensions[col[0].column_letter].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(wrap_text=False, vertical="top")
        ws.freeze_panes = "A2"
    wb.save(path)
    return path


def main() -> None:
    VALIDATION_DIR.mkdir(parents=True, exist_ok=True)

    overlap = build_gene_set_overlap()
    global_tests, pairwise_tests = build_singlecell_tests()
    build_integrated_figure4_anchor()

    epi_means = pd.read_csv(SC_DIR / "GSE116222_signature_gene_state_means.tsv", sep="\t")
    immune_means = pd.read_csv(SC_DIR / "GSE125527_rectal_immune_signature_gene_means.tsv", sep="\t")
    epi_means.to_csv(VALIDATION_DIR / "GSE116222_signature_gene_state_means.tsv", sep="\t", index=False)
    immune_means.to_csv(VALIDATION_DIR / "GSE125527_rectal_immune_signature_gene_means.tsv", sep="\t", index=False)

    build_heatmap_figure(epi_means, immune_means)
    workbook_path = write_workbook(overlap, global_tests, pairwise_tests, epi_means, immune_means)
    print(f"[ok] wrote supplementary validation outputs to {VALIDATION_DIR}")
    print(f"[ok] wrote {workbook_path}")
    print(f"[ok] wrote FigureS3 single-cell heatmaps to {PLOT_DIR}")


if __name__ == "__main__":
    main()
