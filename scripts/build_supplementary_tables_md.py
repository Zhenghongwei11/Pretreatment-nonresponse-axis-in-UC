#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import math
import re
from pathlib import Path


def read_lines(path: Path) -> list[str]:
    return [ln.strip() for ln in path.read_text(encoding="utf-8").splitlines() if ln.strip()]


def _is_number(x: str) -> bool:
    try:
        float(x)
        return True
    except Exception:
        return False


def fmt_float(x: str, ndigits: int) -> str:
    if x in ("", "NA"):
        return x
    try:
        v = float(x)
    except Exception:
        return x
    if math.isnan(v):
        return "NA"
    return f"{v:.{ndigits}f}"


def fmt_p(x: str, sig: int = 3) -> str:
    if x in ("", "NA"):
        return x
    try:
        v = float(x)
    except Exception:
        return x
    if math.isnan(v):
        return "NA"
    if v == 0.0:
        return "0"
    # For readability (avoid "engineering" precision):
    # - Use scientific notation for very small values (< 1e-3), with `sig` significant digits.
    # - Otherwise, cap at 3 decimals (and strip trailing zeros).
    if v < 0.001:
        # `sig` significant digits => `sig-1` digits after the decimal in mantissa.
        s = f"{v:.{max(sig - 1, 0)}e}"
        # Normalize exponent (e-04 -> e-4) for cleaner journal-style display.
        s = re.sub(r"e([+-])0*(\d+)", r"e\1\2", s)
        return s
    return f"{v:.3f}".rstrip("0").rstrip(".")


def fmt_leading_edge(x: str, max_genes: int = 20) -> str:
    if x in ("", "NA"):
        return x
    genes = [g for g in x.split("|") if g]
    if not genes:
        return ""
    if len(genes) <= max_genes:
        return ", ".join(genes)
    return ", ".join(genes[:max_genes]) + ", ..."


def esc_md(s: str) -> str:
    s = s.replace("\r", " ").replace("\n", " ")
    return s.replace("|", "\\|")


def read_tsv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        header = reader.fieldnames or []
        rows = list(reader)
    return header, rows


def read_xlsx(path: Path, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, str]]]:
    # Keep dependency lightweight and local to this function.
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_tuple = next(rows_iter, None)
    if not header_tuple:
        return [], []
    header = [str(x) if x is not None else "" for x in header_tuple]
    out_rows: list[dict[str, str]] = []
    for r in rows_iter:
        if r is None:
            continue
        rr: dict[str, str] = {}
        empty = True
        for h, v in zip(header, r):
            if v is None:
                rr[h] = ""
                continue
            empty = False
            rr[h] = str(v)
        if not empty:
            out_rows.append(rr)
    return header, out_rows


def to_md_table(header: list[str], rows: list[dict[str, str]]) -> str:
    out: list[str] = []
    out.append("| " + " | ".join(esc_md(h) for h in header) + " |")
    out.append("| " + " | ".join(["---"] * len(header)) + " |")
    for r in rows:
        out.append("| " + " | ".join(esc_md(r.get(h, "")) for h in header) + " |")
    return "\n".join(out)


def format_table_s2(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    out = []
    for r in rows:
        rr = dict(r)
        for k in ("Effect (Hedges g)", "Lower 95% CI", "Upper 95% CI", "Adjusted effect (inflammatory proxy)"):
            if k in rr:
                rr[k] = fmt_float(rr[k], 2)
        for k in ("P-value", "FDR", "Adjusted P-value (inflammatory proxy)", "Adjusted FDR (inflammatory proxy)"):
            if k in rr:
                rr[k] = fmt_p(rr[k], sig=3)
        out.append(rr)
    return out


def format_table_s3(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    out = []
    for r in rows:
        rr = dict(r)
        for k in ("Base effect (Hedges g)", "Adjusted effect (Hedges g)", "Lower 95% CI", "Upper 95% CI"):
            if k in rr:
                rr[k] = fmt_float(rr[k], 2)
        for k in ("P-value", "FDR"):
            if k in rr:
                rr[k] = fmt_p(rr[k], sig=3)
        out.append(rr)
    return out


def format_table_s4(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    out = []
    for r in rows:
        rr = dict(r)
        for k in ("P-value", "FDR"):
            if k in rr:
                rr[k] = fmt_p(rr[k], sig=3)
        for k in ("ES", "NES", "log2(error)"):
            if k in rr:
                rr[k] = fmt_float(rr[k], 2)
        # `leadingEdge` lists can be extremely long; keep DOCX-readable.
        if "Leading edge genes" in rr:
            rr["Leading edge genes"] = fmt_leading_edge(rr["Leading edge genes"], max_genes=20)
        # Keep size as integer if possible.
        if "Gene set size" in rr and _is_number(rr["Gene set size"]):
            rr["Gene set size"] = str(int(float(rr["Gene set size"])))
        out.append(rr)
    return out


def format_table_s5(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    out = []
    for r in rows:
        rr = dict(r)
        for k in ("Enrichment score", "Median signature score"):
            if k in rr:
                rr[k] = fmt_float(rr[k], 3)
        for k in ("Cells (n)", "Present signature genes (n)"):
            if k in rr and _is_number(rr[k]):
                rr[k] = str(int(float(rr[k])))
        out.append(rr)
    return out


def format_table_s6(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    out = []
    for r in rows:
        rr = dict(r)
        for k in ("Hedges g (nonresponder - responder)", "Lower 95% CI", "Upper 95% CI"):
            if k in rr:
                rr[k] = fmt_float(rr[k], 2)
        for k in ("P-value (Welch t-test)",):
            if k in rr:
                rr[k] = fmt_p(rr[k], sig=3)
        for k in ("N total", "N responders", "N nonresponders", "Genes used (n)"):
            if k in rr and _is_number(rr[k]):
                rr[k] = str(int(float(rr[k])))
        out.append(rr)
    return out



def format_sheet_gene_set_overlap(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    out = []
    for r in rows:
        rr = dict(r)
        for k in ("fixed_genes_present_in_lodo_pct", "lodo_genes_present_in_fixed_pct"):
            if k in rr and rr[k]:
                try:
                    pct = float(rr[k]) * 100
                    rr[k] = f"{pct:.1f}%"
                except ValueError:
                    pass
        if "jaccard_index" in rr and rr["jaccard_index"]:
            rr["jaccard_index"] = fmt_float(rr["jaccard_index"], 3)
        if "overlap_genes" in rr:
            rr["overlap_genes"] = fmt_leading_edge(rr["overlap_genes"], max_genes=20)
        out.append(rr)
    return out


def format_sheet_sc_global_tests(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    out = []
    for r in rows:
        rr = dict(r)
        if "statistic" in rr and rr["statistic"]:
            rr["statistic"] = fmt_float(rr["statistic"], 2)
        if "pvalue" in rr and rr["pvalue"]:
            rr["pvalue"] = fmt_p(rr["pvalue"], sig=3)
        out.append(rr)
    return out


def format_sheet_sc_pairwise_tests(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    out = []
    for r in rows:
        rr = dict(r)
        for k in ("group_mean", "comparator_mean"):
            if k in rr and rr[k]:
                rr[k] = fmt_float(rr[k], 4)
        if "statistic" in rr and rr["statistic"]:
            try:
                rr["statistic"] = str(int(float(rr["statistic"])))
            except ValueError:
                pass
        for k in ("pvalue", "p_adj_bh"):
            if k in rr and rr[k]:
                rr[k] = fmt_p(rr[k], sig=3)
        out.append(rr)
    return out


def format_sheet_gene_means(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    out = []
    for r in rows:
        rr = dict(r)
        for k, v in rr.items():
            if k != "gene" and v:
                rr[k] = fmt_float(v, 4)
        out.append(rr)
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--submission-dir", default=".", help="Directory containing supplementary_tables/.")
    args = ap.parse_args()

    root = Path(args.submission_dir)
    sup_dir = root / "supplementary_tables"

    # Prefer reviewer-facing XLSX exports if present; fall back to TSV/TXT if needed.
    # Prefer the reviewer-facing filename (but accept legacy names too).
    s1_xlsx = sup_dir / "TableS1_68gene_consensus_signature.xlsx"
    s1_xlsx_legacy = sup_dir / "TableS1_strict_nonresponse_module_genes.xlsx"
    s1_txt = sup_dir / "TableS1_strict_nonresponse_module_genes.txt"
    if s1_xlsx.exists():
        s1_h, s1_rows = read_xlsx(s1_xlsx)
        s1_lines = to_md_table(s1_h, s1_rows)
    elif s1_xlsx_legacy.exists():
        s1_h, s1_rows = read_xlsx(s1_xlsx_legacy)
        s1_lines = to_md_table(s1_h, s1_rows)
    else:
        s1_genes = read_lines(s1_txt)
        s1_lines_list = ["| Index | Gene symbol |", "| ---: | --- |"]
        for i, g in enumerate(s1_genes, 1):
            s1_lines_list.append(f"| {i} | {esc_md(g)} |")
        s1_lines = "\n".join(s1_lines_list)

    def read_table(name: str) -> tuple[list[str], list[dict[str, str]]]:
        xlsx = sup_dir / f"{name}.xlsx"
        tsv = sup_dir / f"{name}.tsv"
        if xlsx.exists():
            return read_xlsx(xlsx)
        return read_tsv(tsv)

    s2_h, s2 = read_table("TableS2_replication_effect_sizes")
    s3_h, s3 = read_table("TableS3_specificity_checks")
    s4_h, s4 = read_table("TableS4_hallmark_fgsea_all")
    s5_h, s5 = read_table("TableS5_singlecell_localization")
    s6_path_xlsx = sup_dir / "TableS6_gse92415_antiTNF_class_sensitivity.xlsx"
    s6_path_tsv = sup_dir / "TableS6_gse92415_antiTNF_class_sensitivity.tsv"
    s6_h: list[str] = []
    s6: list[dict[str, str]] = []
    if s6_path_xlsx.exists() or s6_path_tsv.exists():
        s6_h, s6 = read_table("TableS6_gse92415_antiTNF_class_sensitivity")

    s2_md = to_md_table(s2_h, format_table_s2(s2))
    s3_md = to_md_table(s3_h, format_table_s3(s3))
    s4_md = to_md_table(s4_h, format_table_s4(s4))
    s5_md = to_md_table(s5_h, format_table_s5(s5))
    s6_md = to_md_table(s6_h, format_table_s6(s6)) if s6_h else ""

    md: list[str] = []
    md.append("# Supplementary Tables")
    md.append("")
    md.append("This document contains Tables S1–S7 for separate upload.")
    md.append("")

    md.append("## Table S1. Strict 68-gene cross-cohort consensus signature (gene symbols)")
    md.append("Gene symbols for the strict consensus signature used for interpretability, specificity benchmarking, and single-cell localization.")
    md.append("")
    md.append(s1_lines)
    md.append("")

    md.append("## Table S2. Replication effect sizes (consensus signature and LODO axis)")
    md.append("Numeric fields are rounded for readability; full-precision values are available in the corresponding TSV files.")
    md.append("")
    md.append(s2_md)
    md.append("")

    md.append("## Table S3. Specificity checks under inflammation and remodeling proxy adjustment")
    md.append("Numeric fields are rounded for readability; full-precision values are available in the corresponding TSV files.")
    md.append("")
    md.append(s3_md)
    md.append("")

    md.append("## Table S4. Hallmark pathway enrichment results (full FGSEA outputs)")
    md.append("Numeric fields are rounded for readability. `leadingEdge` is truncated in this DOCX-friendly rendering; full outputs (including full `leadingEdge`) are available in the corresponding TSV files.")
    md.append("")
    md.append(s4_md)
    md.append("")

    md.append("## Table S5. Single-cell localization support summary (processed public references)")
    md.append("Numeric fields are rounded for readability; full-precision values are available in the corresponding TSV files.")
    md.append("")
    md.append(s5_md)
    md.append("")

    if s6_md:
        md.append("## Table S6. Class-level anti-TNF sensitivity analysis (GSE92415, golimumab-treated UC cohort)")
        md.append("This table reports the association between the fixed 68-gene consensus signature and Week 6 response status in the golimumab-treated UC cohort.")
        md.append("")
        md.append(s6_md)
        md.append("")

    s7_path = sup_dir / "TableS7_gene_set_overlap_and_singlecell_tests.xlsx"
    if s7_path.exists():
        md.append("## Table S7. Gene-set overlap and single-cell statistical tests")
        md.append("This table reports the gene-set overlap between the fixed consensus signature and the broader LODO training-derived axes, as well as global and pairwise single-cell statistical tests and cell-type gene means for the consensus signature.")
        md.append("")

        # Sheet 1: GeneSetOverlap
        s7_1_h, s7_1_r = read_xlsx(s7_path, "GeneSetOverlap")
        s7_1_md = to_md_table(s7_1_h, format_sheet_gene_set_overlap(s7_1_r))
        md.append("### Sheet 1: GeneSetOverlap")
        md.append("This sheet reports the overlap between the fixed consensus signature (68 genes) and the broader LODO training-derived axes (held-out datasets).")
        md.append("")
        md.append(s7_1_md)
        md.append("")

        # Sheet 2: SingleCellGlobalTests
        s7_2_h, s7_2_r = read_xlsx(s7_path, "SingleCellGlobalTests")
        s7_2_md = to_md_table(s7_2_h, format_sheet_sc_global_tests(s7_2_r))
        md.append("### Sheet 2: SingleCellGlobalTests")
        md.append("This sheet reports global Kruskal-Wallis test statistics and p-values for signature activity differences across cell types or states.")
        md.append("")
        md.append(s7_2_md)
        md.append("")

        # Sheet 3: SingleCellPairwiseTests
        s7_3_h, s7_3_r = read_xlsx(s7_path, "SingleCellPairwiseTests")
        s7_3_md = to_md_table(s7_3_h, format_sheet_sc_pairwise_tests(s7_3_r))
        md.append("### Sheet 3: SingleCellPairwiseTests")
        md.append("This sheet reports pairwise Mann-Whitney U test statistics, direction, and adjusted p-values (Benjamini-Hochberg) for signature activity differences.")
        md.append("")
        md.append(s7_3_md)
        md.append("")

        # Sheet 4: EpithelialGeneMeans
        s7_4_h, s7_4_r = read_xlsx(s7_path, "EpithelialGeneMeans")
        s7_4_md = to_md_table(s7_4_h, format_sheet_gene_means(s7_4_r))
        md.append("### Sheet 4: EpithelialGeneMeans")
        md.append("This sheet reports the mean expression scores of the consensus signature genes across marker-defined epithelial states.")
        md.append("")
        md.append(s7_4_md)
        md.append("")

        # Sheet 5: ImmuneGeneMeans
        s7_5_h, s7_5_r = read_xlsx(s7_path, "ImmuneGeneMeans")
        s7_5_md = to_md_table(s7_5_h, format_sheet_gene_means(s7_5_r))
        md.append("### Sheet 5: ImmuneGeneMeans")
        md.append("This sheet reports the mean expression scores of the consensus signature genes across marker-defined immune cell types in diseased and healthy rectal tissues.")
        md.append("")
        md.append(s7_5_md)
        md.append("")

    out_path = root / "SUPPLEMENTARY_TABLES.md"
    out_path.write_text("\n".join(md), encoding="utf-8")
    print(f"[ok] wrote {out_path}")


if __name__ == "__main__":
    main()
