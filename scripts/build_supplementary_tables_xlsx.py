#!/usr/bin/env python3
from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
SUP_DIR = ROOT / "supplementary_tables"
OUT_PATH = SUP_DIR / "SUPPLEMENTARY_TABLES.xlsx"


TABLE_FILES = [
    ("S0_Index", "TableS0_INDEX.xlsx", None),
    ("S1_68GeneSignature", "TableS1_68gene_consensus_signature.xlsx", None),
    ("S2_EffectSizes", "TableS2_replication_effect_sizes.xlsx", None),
    ("S3_SpecificityChecks", "TableS3_specificity_checks.xlsx", None),
    ("S4_HallmarkFGSEA", "TableS4_hallmark_fgsea_all.xlsx", None),
    ("S5_SingleCellSummary", "TableS5_singlecell_localization.xlsx", None),
    ("S6_GSE92415", "TableS6_gse92415_antiTNF_class_sensitivity.xlsx", None),
    ("S7_GeneSetOverlap", "TableS7_gene_set_overlap_and_singlecell_tests.xlsx", "GeneSetOverlap"),
    ("S7_SCGlobalTests", "TableS7_gene_set_overlap_and_singlecell_tests.xlsx", "SingleCellGlobalTests"),
    ("S7_SCPairwiseTests", "TableS7_gene_set_overlap_and_singlecell_tests.xlsx", "SingleCellPairwiseTests"),
    ("S7_EpithelialMeans", "TableS7_gene_set_overlap_and_singlecell_tests.xlsx", "EpithelialGeneMeans"),
    ("S7_ImmuneMeans", "TableS7_gene_set_overlap_and_singlecell_tests.xlsx", "ImmuneGeneMeans"),
]


def copy_sheet(source_path: Path, source_sheet: str | None, target_wb: Workbook, target_name: str) -> None:
    source_wb = load_workbook(source_path, data_only=False)
    source_ws = source_wb[source_sheet] if source_sheet else source_wb.worksheets[0]
    target_ws = target_wb.create_sheet(target_name)

    for row in source_ws.iter_rows():
        for cell in row:
            out_cell = target_ws[cell.coordinate]
            out_cell.value = cell.value
            if cell.has_style:
                out_cell.font = copy(cell.font)
                out_cell.fill = copy(cell.fill)
                out_cell.border = copy(cell.border)
                out_cell.alignment = copy(cell.alignment)
                out_cell.number_format = cell.number_format
                out_cell.protection = copy(cell.protection)
            if cell.hyperlink:
                out_cell._hyperlink = copy(cell.hyperlink)
            if cell.comment:
                out_cell.comment = copy(cell.comment)

    for key, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[key].width = dim.width
    for key, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[key].height = dim.height
    target_ws.freeze_panes = source_ws.freeze_panes


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for target_name, filename, source_sheet in TABLE_FILES:
        source_path = SUP_DIR / filename
        if not source_path.exists():
            raise FileNotFoundError(source_path)
        copy_sheet(source_path, source_sheet, wb, target_name)
    wb.save(OUT_PATH)
    print(f"[ok] wrote {OUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
